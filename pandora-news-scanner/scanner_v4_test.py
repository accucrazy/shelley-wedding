# -*- coding: utf-8 -*-
"""
Pandora Scanner v4.1 — CLI 介面 + 日期驗證 + 三層搜尋 + 強化 LLM

用法:
  python scanner_v4_test.py scan [--task 全家草莓季] [--month 一月]
  python scanner_v4_test.py report [--vs-old] [--vs-yilan]
  python scanner_v4_test.py press [--list] [--add]
  python scanner_v4_test.py results [--task 全家草莓季] [--check-quality]
"""

import json, os, sys, time, re, argparse
from datetime import datetime, date
from urllib.parse import urlparse, quote
from pathlib import Path
from xml.etree import ElementTree as ET
from collections import Counter

sys.path.insert(0, os.path.expanduser(
    "~/openclaw/skills/pandora-news/venv/lib/python3.9/site-packages"))

import httpx
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich.layout import Layout
from rich.text import Text
from rich import box

try:
    from ddgs import DDGS
except ImportError:
    DDGS = None

if sys.stdout:
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

console = Console()

# ═══════════════════════════════════════════════════════
# Paths
# ═══════════════════════════════════════════════════════

BASE_DIR = Path(os.path.expanduser(
    "~/openclaw/workspace/newscollect/Pandora News IO 0226"))
MEDIA_LIST = BASE_DIR / "1-媒體清單" / "news-media-list.md"
OUTPUT_DIR = Path(__file__).parent / "v4_test_output"
OLD_RESULTS_DIR = BASE_DIR / "3-Pandora News io" / "一月"
YILAN_DIR = Path(os.path.expanduser(
    "~/Downloads/Pandora News IO 0226/0-意藍新聞IO/1月"))

SEARCH_DELAY = 1.5
DDG_MAX_RETRIES = 3

AGGREGATOR_DOMAINS = {
    "today.line.me", "tw.news.yahoo.com", "msn.com", "news.pchome.com.tw"}

# ═══════════════════════════════════════════════════════
# Task config — 每個任務附帶新聞稿發稿日
# ═══════════════════════════════════════════════════════

TASKS_CONFIG = {
    # ─── 全家 一月 ───
    "全家草莓季": {
        "keywords": ["全家草莓季", "全家 ASAMIMICHAN", "全家 草莓霜淇淋",
                     "全家 莓好運輸中", "全家 草莓優格霜淇淋", "FamilyMart 草莓季"],
        "press_title_fragments": [
            "草莓季17粉嫩登場",
            "ASAMIMICHAN萌翻全台",
            "草莓優格霜淇淋 草莓厚奶雲餡泡芙",
        ],
        "brand_keywords": ["草莓", "莓", "ASAMIMICHAN", "全家"],
        "press_date": "2026-01-05",
        "month": "一月",
        "brand": "全家",
    },
    "全家UCC咖啡": {
        "keywords": ["全家 UCC咖啡", "全家 阿里山極選", "全家 Let's Café 阿里山",
                     "全家 雙冠軍監製", "全家 阿里山極選綜合咖啡"],
        "press_title_fragments": ["再攜UCC推雙冠軍監製", "阿里山極選綜合咖啡65元"],
        "brand_keywords": ["UCC", "阿里山", "咖啡", "全家"],
        "press_date": "2026-01-21",
        "month": "一月",
        "brand": "全家",
    },
    "全家開運鮮食": {
        "keywords": ["全家 開運鮮食", "全家 紅運烏魚子", "全家 蘭州拉麵",
                     "全家 開運鮮食祭"],
        "press_title_fragments": ["開運鮮食祭", "紅運烏魚子"],
        "brand_keywords": ["開運", "鮮食", "烏魚子", "全家"],
        "press_date": "2026-01-20",
        "month": "一月",
        "brand": "全家",
    },
    "全家年菜預購": {
        "keywords": ["全家 年菜預購", "全家 2026金馬年菜", "全家 FamiPort 年菜",
                     "全家 星級年菜", "全家 富錦樹 年菜"],
        "press_title_fragments": ["搶攻圍爐商機 2026金馬年菜", "FamiPort一站購足 星級名店"],
        "brand_keywords": ["年菜", "預購", "圍爐", "全家"],
        "press_date": "2026-01-14",
        "month": "一月",
        "brand": "全家",
    },
    "全家蜷川實花": {
        "keywords": ["全家 蜷川實花", "全家 蜷川實花展", "全家 蜷川實花 杯身"],
        "press_title_fragments": ["攜手蜷川實花展推獨家限定杯身", "全家 蜷川實花 杯套"],
        "brand_keywords": ["蜷川", "實花", "Ninagawa", "全家"],
        "press_date": "2026-01-12",
        "month": "一月",
        "brand": "全家",
    },
    "全家特力屋": {
        "keywords": ["全家 特力屋", "全家 居家微整型", "全家 免治馬桶 特力屋",
                     "全家行動購 特力屋"],
        "press_title_fragments": ["攜特力屋齊推居家微整型", "免治馬桶 電子鎖8990元含安裝"],
        "brand_keywords": ["特力屋", "居家", "免治", "全家"],
        "press_date": "2026-01-06",
        "month": "一月",
        "brand": "全家",
    },
    "全家溏心蛋": {
        "keywords": ["全家 溏心蛋", "全家 日式溏心蛋", "全家 用撈的 溏心蛋",
                     "全家 溏心蛋 25元"],
        "press_title_fragments": ["首推用撈的日式溏心蛋", "熟食區新蛋報到 溏心蛋25元"],
        "brand_keywords": ["溏心蛋", "日式", "全家"],
        "press_date": "2026-01-28",
        "month": "一月",
        "brand": "全家",
    },
    "全家高山茶": {
        "keywords": ["全家 高山茶", "全家 蘭韻梨山烏龍", "全家 Let's Tea 高山茶",
                     "全家 現煮精品茶"],
        "press_title_fragments": ["寒流飄茶香 高山茶進駐全家", "蘭韻梨山烏龍49元"],
        "brand_keywords": ["高山茶", "梨山", "烏龍", "全家"],
        "press_date": "2026-01-08",
        "month": "一月",
        "brand": "全家",
    },
    "全家超人力霸王": {
        "keywords": ["全家 超人力霸王", "全家 高雄冬日遊樂園", "全家 超人力霸王 聯名",
                     "全家 超人力霸王 杯塞"],
        "press_title_fragments": ["超人力霸王60周年降臨高雄", "全家 超人力霸王 聯名杯塞 拍拍燈"],
        "brand_keywords": ["超人力霸王", "Ultraman", "全家"],
        "press_date": "2026-01-19",
        "month": "一月",
        "brand": "全家",
    },
    "全家寒流抗寒": {
        "keywords": ["全家 寒流", "全家 抗寒", "全家 寒流 熱食", "全家 暖暖包",
                     "全家 寒流 熱飲"],
        "press_title_fragments": ["強烈冷氣團 全家 熱飲熱食", "全家 抗寒限定好康"],
        "brand_keywords": ["寒流", "抗寒", "暖", "全家"],
        "press_date": "2026-01-19",
        "month": "一月",
        "brand": "全家",
    },
    "全家伴手禮": {
        "keywords": ["全家 陳耀訓", "全家 紅土蛋黃酥", "全家 伴手禮 過年",
                     "全家 春節伴手禮"],
        "press_title_fragments": ["陳耀訓 紅土蛋黃酥"],
        "brand_keywords": ["陳耀訓", "蛋黃酥", "伴手禮", "全家"],
        "press_date": "2026-01-30",
        "month": "一月",
        "brand": "全家",
    },

    # ─── 全家 二月 ───
    "全家減糖無糖飲品": {
        "keywords": [
            "全家 國健署 減糖",
            "全家 營養師公會 無糖飲品",
            "全家 減糖指標 推廣",
            "FamilyMart 無糖飲品",
            "全家 響應 國健署 減糖",
            "全家 營養師 無糖 推廣",
            "全家 無糖茶 無糖咖啡 推廣",
            "全家便利商店 減糖 國健署",
            "全家 營養師公會全國聯合會",
        ],
        "press_title_fragments": [
            "響應國健署減糖指標",
            "攜手營養師公會全國聯合會",
            "推廣無糖飲品",
        ],
        "brand_keywords": ["減糖", "無糖", "國健署", "營養師公會", "全家"],
        "press_date": "2026-02-12",
        "month": "二月",
        "brand": "全家",
    },

    # ─── 7-ELEVEN 一月 ───
    "711清水服務區週年慶": {
        "keywords": ["統一超商 清水服務區", "7-ELEVEN 清水服務區", "清水服務區 週年慶",
                     "清水服務區 懂買懂吃懂拍", "7-11 清水服務區",
                     "清水服務區 一週年", "統一超商 清水 旅遊攻略",
                     "國道三號 清水服務區", "清水休息站 週年"],
        "press_title_fragments": ["清水服務區1週年慶", "懂買、懂吃、懂拍旅遊攻略",
                                  "統一超商經營清水服務區"],
        "brand_keywords": ["清水服務區", "週年慶", "7-ELEVEN", "統一超商"],
        "press_date": "2026-01-02",
        "month": "一月",
        "brand": "7-11",
    },
    "711米其林法餐": {
        "keywords": ["7-ELEVEN 米其林 法餐", "7-ELEVEN 30顆星主廚", "7-ELEVEN 低食物里程燉飯",
                     "7-11 米其林 主廚監製", "7-ELEVEN 星級儀式感"],
        "press_title_fragments": ["米其林30顆星主廚監製法餐上桌", "低食物里程燉飯"],
        "brand_keywords": ["米其林", "法餐", "燉飯", "7-ELEVEN"],
        "press_date": "2026-01-06",
        "month": "一月",
        "brand": "7-11",
    },
    "711年節社交經濟": {
        "keywords": ["7-ELEVEN 年節 社交經濟", "7-ELEVEN 開運年菜 尾牙家電",
                     "7-ELEVEN 年貨大街", "7-11 年菜 尾牙", "7-ELEVEN i預購 年節",
                     "統一超商 年節 年菜", "7-ELEVEN i預購 年菜 2026",
                     "7-ELEVEN 尾牙 家電 獎品", "統一超商 年貨 預購",
                     "7-ELEVEN 社交經濟"],
        "press_title_fragments": ["搶攻年節「社交經濟」", "開運年菜、尾牙家電馬上登場",
                                  "7-ELEVEN全面搶攻年節"],
        "brand_keywords": ["年節", "年菜", "尾牙", "年貨", "社交經濟", "i預購", "7-ELEVEN"],
        "press_date": "2026-01-07",
        "month": "一月",
        "brand": "7-11",
    },
    "711把愛找回來公益": {
        "keywords": ["7-ELEVEN 把愛找回來", "7-ELEVEN 公益募款", "7-ELEVEN 130個公益團體",
                     "7-11 把愛找回來 2026", "7-ELEVEN 暖心公益"],
        "press_title_fragments": ["把愛找回來公益募款平台", "擴大合作130個公益團體"],
        "brand_keywords": ["把愛找回來", "公益", "募款", "7-ELEVEN"],
        "press_date": "2026-01-08",
        "month": "一月",
        "brand": "7-11",
    },
    "711智能果昔機": {
        "keywords": ["7-ELEVEN 智能果昔機", "7-ELEVEN 鮮玉米濃湯", "7-11 果昔機",
                     "7-ELEVEN 現打果昔", "7-11 智能果昔機 鮮玉米",
                     "統一超商 果昔機", "7-ELEVEN 果昔 鮮榨 2026",
                     "7-ELEVEN 智能 現打 玉米", "統一超商 鮮玉米濃湯"],
        "press_title_fragments": ["首度導入「智能果昔機」", "鮮玉米濃湯 現打熱熱喝",
                                  "7-ELEVEN首度導入"],
        "brand_keywords": ["果昔機", "鮮玉米", "現打", "智能", "7-ELEVEN"],
        "press_date": "2026-01-09",
        "month": "一月",
        "brand": "7-11",
    },
    "711阜杭豆漿飯糰": {
        "keywords": ["7-ELEVEN 阜杭豆漿", "7-ELEVEN 阜杭豆漿飯糰", "7-11 阜杭豆漿",
                     "7-ELEVEN 學測 開運", "7-ELEVEN 神明加持好運"],
        "press_title_fragments": ["阜杭豆漿飯糰 締造億元", "學測考生必吃"],
        "brand_keywords": ["阜杭豆漿", "飯糰", "學測", "7-ELEVEN"],
        "press_date": "2026-01-12",
        "month": "一月",
        "brand": "7-11",
    },
    "711小熊維尼集點": {
        "keywords": ["7-ELEVEN 小熊維尼", "7-ELEVEN 維尼100周年", "7-ELEVEN 全店集點 維尼",
                     "7-11 小熊維尼 集點", "7-ELEVEN 魯斯佛"],
        "press_title_fragments": ["小熊維尼滿100周年", "80款獨家療癒新品", "魯斯佛"],
        "brand_keywords": ["小熊維尼", "維尼", "集點", "魯斯佛", "7-ELEVEN"],
        "press_date": "2026-01-13",
        "month": "一月",
        "brand": "7-11",
    },
    "711馬年前哨戰優惠": {
        "keywords": ["7-ELEVEN 馬年 優惠", "7-ELEVEN 金馬年 前哨戰",
                     "7-11 CITY TEA 優惠", "7-ELEVEN 甜甜圈炸雞堡",
                     "統一超商 馬年 前哨戰", "7-ELEVEN 0115 優惠",
                     "7-11 金馬年 甜甜圈", "7-ELEVEN 1月15日 優惠",
                     "7-ELEVEN 馬年前哨戰 特價", "統一超商 金馬 優惠"],
        "press_title_fragments": ["馬年前哨戰", "甜甜圈炸雞堡", "0115優惠資訊"],
        "brand_keywords": ["馬年", "優惠", "前哨戰", "甜甜圈", "7-ELEVEN"],
        "press_date": "2026-01-15",
        "month": "一月",
        "brand": "7-11",
    },
    "711Fresh橋港門市": {
        "keywords": ["7-ELEVEN Fresh橋港門市", "7-ELEVEN 八里淡水 Fresh", "7-11 橋港門市",
                     "7-ELEVEN 淡江大橋 門市", "7-ELEVEN 雙北首間Fresh"],
        "press_title_fragments": ["Fresh橋港門市", "卡位八里淡水生活圈"],
        "brand_keywords": ["Fresh", "橋港", "淡江大橋", "7-ELEVEN"],
        "press_date": "2026-01-15",
        "month": "一月",
        "brand": "7-11",
    },
    "711金馬開運活動": {
        "keywords": ["7-ELEVEN 金馬開運好運來", "7-ELEVEN 金馬開運 活動",
                     "7-11 滿額立折 抽抽樂", "7-ELEVEN 滿222元 折扣抽抽樂",
                     "7-ELEVEN 金馬開運主題專案架"],
        "press_title_fragments": ["金馬開運好運來", "滿額立折抽抽樂"],
        "brand_keywords": ["金馬", "開運", "抽抽樂", "滿額", "7-ELEVEN"],
        "press_date": "2026-01-15",
        "month": "一月",
        "brand": "7-11",
    },
    "711藝伎咖啡": {
        "keywords": ["7-ELEVEN 藝伎咖啡", "阿里山豆御香 藝伎", "CITY CAFE 藝伎",
                     "7-11 藝伎咖啡 2000元", "7-ELEVEN 蕉財進寶 香蕉拿鐵",
                     "統一超商 藝伎咖啡", "7-ELEVEN 藝伎 限量",
                     "CITY CAFE 阿里山 藝伎", "7-ELEVEN 咖啡 2000元",
                     "統一超商 藝伎 限定門市", "7-ELEVEN 精品咖啡 藝伎",
                     "阿里山 藝伎 7-ELEVEN", "CITY CAFE 2000元 咖啡"],
        "press_title_fragments": ["阿里山豆御香藝伎咖啡", "每杯2,000元", "蕉財進寶",
                                  "真的不可思議", "限定門市限量開賣"],
        "brand_keywords": ["藝伎", "御香", "蕉財進寶", "CITY CAFE", "7-ELEVEN", "2000元"],
        "press_date": "2026-01-19",
        "month": "一月",
        "brand": "7-11",
    },
    "711抗寒保暖": {
        "keywords": ["7-ELEVEN 抗寒保暖", "7-ELEVEN 抗寒 優惠", "7-11 寒流 暖暖包",
                     "7-ELEVEN 發熱衣 優惠"],
        "press_title_fragments": ["抗寒保暖", "7-ELEVEN 暖暖包"],
        "brand_keywords": ["抗寒", "保暖", "暖暖包", "發熱衣", "7-ELEVEN"],
        "press_date": "2026-01-20",
        "month": "一月",
        "brand": "7-11",
    },
    "711金馬年開運": {
        "keywords": ["7-ELEVEN 馬上開心 金馬年", "7-ELEVEN 開運美食 桂氣茶飲",
                     "7-11 金馬年 多重回饋", "7-ELEVEN 18項開運美食",
                     "統一超商 金馬年 開運", "7-ELEVEN 乖乖 金馬",
                     "7-ELEVEN 馬上開心 開運", "統一超商 馬年 回饋",
                     "7-ELEVEN 乖乖 聯名 金馬"],
        "press_title_fragments": ["馬上開心 7-ELEVEN多重回饋", "桂氣茶飲好運一起發",
                                  "新年到「馬上開心」", "超過18項開運美食"],
        "brand_keywords": ["金馬年", "馬上開心", "開運美食", "桂氣茶飲", "乖乖", "7-ELEVEN"],
        "press_date": "2026-01-20",
        "month": "一月",
        "brand": "7-11",
    },
    "711東南亞美食": {
        "keywords": ["7-ELEVEN 東南亞美食", "7-ELEVEN 道地家鄉味", "7-11 東南亞 零食",
                     "7-ELEVEN 多元共好 消費"],
        "press_title_fragments": ["東南亞美食 來7-ELEVEN", "道地家鄉味"],
        "brand_keywords": ["東南亞", "家鄉味", "多元共好", "7-ELEVEN"],
        "press_date": "2026-01-23",
        "month": "一月",
        "brand": "7-11",
    },
    "711桂氣茶飲優惠": {
        "keywords": ["7-ELEVEN 桂氣茶飲", "7-ELEVEN 開運 手搖控", "7-11 CITY TEA 桂",
                     "7-ELEVEN 優惠 0126",
                     "統一超商 桂氣茶飲", "7-ELEVEN 桂花 茶飲 開運",
                     "CITY TEA 桂花", "7-11 手搖 開運 桂",
                     "7-ELEVEN 1月26日 優惠", "統一超商 手搖控",
                     "7-ELEVEN CITY TEA 桂花 2026"],
        "press_title_fragments": ["開運「桂」氣茶飲", "手搖控開喝",
                                  "0126優惠資訊"],
        "brand_keywords": ["桂氣", "茶飲", "CITY TEA", "桂花", "手搖控", "7-ELEVEN"],
        "press_date": "2026-01-26",
        "month": "一月",
        "brand": "7-11",
    },
    "711西村優志集點": {
        "keywords": ["7-ELEVEN 西村優志", "7-ELEVEN 開心熊貓 小鼠 戀愛兔",
                     "7-11 西村優志 集點", "7-ELEVEN 精品集點 西村",
                     "統一超商 西村優志", "7-ELEVEN 西村優志 Yushi Nishimura",
                     "7-11 精品集點 2026", "7-ELEVEN 開心熊貓",
                     "7-ELEVEN 戀愛兔 集點"],
        "press_title_fragments": ["首度攜手日本人氣創作者「西村優志」", "開心熊貓、小鼠、戀愛兔",
                                  "全店精品集點首度攜手"],
        "brand_keywords": ["西村優志", "開心熊貓", "戀愛兔", "集點", "7-ELEVEN"],
        "press_date": "2026-01-27",
        "month": "一月",
        "brand": "7-11",
    },
    "711開運福袋": {
        "keywords": ["7-ELEVEN 開運福袋", "7-ELEVEN 福袋 第二波", "7-11 福袋 IP肖像",
                     "7-ELEVEN 發財金 福袋", "7-ELEVEN 牡蠣寶寶",
                     "統一超商 福袋 2026", "7-ELEVEN 福袋 限量",
                     "7-11 第二波 福袋 IP", "統一超商 開運福袋",
                     "7-ELEVEN 福袋 2月"],
        "press_title_fragments": ["第二波14款開運福袋", "7大超萌IP肖像",
                                  "2月11日起限量開賣"],
        "brand_keywords": ["福袋", "開運", "牡蠣寶寶", "發財金", "限量", "7-ELEVEN"],
        "press_date": "2026-01-29",
        "month": "一月",
        "brand": "7-11",
    },
    "711香菜美食": {
        "keywords": ["7-ELEVEN 香菜美食", "7-ELEVEN 開香 香菜", "7-11 香菜 20款",
                     "7-ELEVEN 就愛這一味 香菜"],
        "press_title_fragments": ["逾20款香菜美食", "等你來「開香」"],
        "brand_keywords": ["香菜", "開香", "7-ELEVEN"],
        "press_date": "2026-01-29",
        "month": "一月",
        "brand": "7-11",
    },
    "711OPEN家族貼圖": {
        "keywords": ["OPEN家族 貼圖", "OPEN家族 免費貼圖 2026", "7-ELEVEN OPEN 拜年",
                     "7-11 OPEN 福氣跑馬燈籠", "柿柿如意 盒玩公仔",
                     "OPEN小將 貼圖", "OPEN將 免費貼圖",
                     "7-ELEVEN OPEN LINE貼圖", "統一超商 OPEN家族 貼圖",
                     "OPEN醬 貼圖 2026", "7-ELEVEN OPEN 拜年神器"],
        "press_title_fragments": ["OPEN!家族2026年16款全新免費貼圖", "福氣跑馬燈籠",
                                  "社群最新拜年神器"],
        "brand_keywords": ["OPEN", "貼圖", "跑馬燈籠", "OPEN小將", "拜年", "7-ELEVEN"],
        "press_date": "2026-01-29",
        "month": "一月",
        "brand": "7-11",
    },
    "711貓福珊迪": {
        "keywords": ["7-ELEVEN 貓福珊迪", "7-ELEVEN 貓之日 貓福珊迪", "7-11 貓福珊迪",
                     "7-ELEVEN 喵財進寶 提拉米蘇 奶茶",
                     "統一超商 貓福珊迪", "7-ELEVEN mofusand",
                     "7-ELEVEN 貓之日 聯名", "mofusand 7-11",
                     "7-ELEVEN 貓福珊迪 甜蜜賣萌", "統一超商 貓之日",
                     "7-ELEVEN mofusand 聯名 2026"],
        "press_title_fragments": ["貓之日 7-ELEVEN聯名「貓福珊迪」", "喵財進寶提拉米蘇風味奶茶",
                                  "貓福珊迪 甜蜜賣萌"],
        "brand_keywords": ["貓福珊迪", "貓之日", "喵財進寶", "mofusand", "7-ELEVEN"],
        "press_date": "2026-01-30",
        "month": "一月",
        "brand": "7-11",
    },

    # ── 7-ELEVEN 二月 ──
    "711開工狂撒幣": {
        "keywords": [
            "7-ELEVEN 狂撒幣 淘金",
            "7-ELEVEN 開工 優惠 CITY",
            "7-ELEVEN 開學 優惠",
            "7-11 開工 咖啡 5折",
            "7-ELEVEN 5千萬 回饋",
            "7-ELEVEN 年後飲控 早餐",
            "統一超商 開工 狂撒幣",
            "7-ELEVEN CITY 系列 5折起",
            "7-ELEVEN 門市淘金 新年",
            "7-11 多重優惠 總價值 5千萬",
            "統一超商 開工開學 優惠",
            "7-ELEVEN 鮮食組合 年後",
            "7-ELEVEN 迎新年 狂撒幣 淘金",
        ],
        "press_title_fragments": [
            "迎新年狂撒幣快來門市淘金",
            "多重優惠回饋總價值達5千萬",
            "開工開學推CITY系列優惠最殺5折起",
            "年後飲控逾百款超值早餐",
        ],
        "brand_keywords": ["狂撒幣", "淘金", "開工", "開學", "CITY", "5折", "5千萬", "飲控", "7-ELEVEN"],
        "press_date": "2026-02-11",
        "month": "二月",
        "brand": "7-11",
    },
}

# ═══════════════════════════════════════════════════════
# Gemini API
# ═══════════════════════════════════════════════════════

def _load_api_key():
    for p in [Path(os.path.expanduser("~/openclaw/.env")),
              Path(os.path.expanduser("~/.openclaw/.env"))]:
        if p.exists():
            for line in p.read_text().splitlines():
                if line.startswith("GOOGLE_API_KEY="):
                    return line.split("=", 1)[1].strip()
    return os.environ.get("GOOGLE_API_KEY", "")

GEMINI_API_KEY = _load_api_key()
GEMINI_URL = ("https://generativelanguage.googleapis.com/v1beta"
              "/models/gemini-2.0-flash:generateContent")

# ═══════════════════════════════════════════════════════
# Media list helpers
# ═══════════════════════════════════════════════════════

def load_media_list():
    content = MEDIA_LIST.read_text(encoding="utf-8")
    media = []
    for line in content.split("\n"):
        if "|" not in line or "---" in line or "媒體名稱" in line:
            continue
        parts = [p.strip() for p in line.split("|") if p.strip()]
        if len(parts) >= 2 and "." in parts[1]:
            media.append({"name": parts[0], "domain": parts[1]})
    return media

def build_domain_map(media_list):
    dmap = {}
    for m in media_list:
        d = m["domain"].replace("www.", "")
        dmap[d] = m["name"]
    for m in media_list:
        d = m["domain"].replace("www.", "")
        parts = d.split(".")
        if len(parts) >= 3:
            short = ".".join(parts[-2:])
            if short not in dmap:
                dmap[short] = m["name"]
    return dmap

def match_domain(url, domain_map):
    try:
        host = urlparse(url).netloc.replace("www.", "").lower()
    except Exception:
        return None, None
    if host in domain_map:
        return domain_map[host], host
    parts = host.split(".")
    for i in range(len(parts)):
        sub = ".".join(parts[i:])
        if sub in domain_map:
            return domain_map[sub], sub
    return None, None

# ═══════════════════════════════════════════════════════
# Date helpers
# ═══════════════════════════════════════════════════════

def parse_date_str(s):
    """Try to parse various date formats, return 'YYYY-MM-DD' or ''."""
    if not s:
        return ""
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    m = re.search(r'(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return ""

def is_after_press_date(article_date, press_date_str):
    """Check if article date is on or after press release date.
    If article has no date, return None (uncertain)."""
    if not article_date:
        return None
    try:
        a = datetime.strptime(article_date[:10], "%Y-%m-%d").date()
        p = datetime.strptime(press_date_str, "%Y-%m-%d").date()
        return a >= p
    except Exception:
        return None

def extract_date_from_snippet(body):
    """Try to extract a date from DDG snippet text."""
    if not body:
        return ""
    patterns = [
        r'(\d{4})[/.-](\d{1,2})[/.-](\d{1,2})',
        r'(\d{4})年(\d{1,2})月(\d{1,2})日',
    ]
    for pat in patterns:
        m = re.search(pat, body)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if 2020 <= y <= 2030 and 1 <= mo <= 12 and 1 <= d <= 31:
                return f"{y}-{mo:02d}-{d:02d}"
    return ""

# ═══════════════════════════════════════════════════════
# Search sources
# ═══════════════════════════════════════════════════════

def google_news_rss(keyword, after_date=None):
    """Search Google News RSS. Optionally filter by date via 'after:' param."""
    q = keyword
    if after_date:
        q = f"{keyword} after:{after_date}"
    url = (f"https://news.google.com/rss/search?"
           f"q={quote(q)}&hl=zh-TW&gl=TW&ceid=TW:zh-Hant")
    try:
        resp = httpx.get(url, timeout=15, follow_redirects=True,
                        headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
    except Exception as e:
        return []
    results = []
    try:
        root = ET.fromstring(resp.text)
        for item in root.iter("item"):
            title = item.findtext("title", "")
            link = item.findtext("link", "")
            pub_date = item.findtext("pubDate", "")
            source_el = item.find("source")
            source_name = item.findtext("source", "")
            source_url = source_el.get("url", "") if source_el is not None else ""
            date_str = ""
            if pub_date:
                try:
                    from email.utils import parsedate_to_datetime
                    dt = parsedate_to_datetime(pub_date)
                    date_str = dt.strftime("%Y-%m-%d")
                except Exception:
                    pass
            results.append({
                "title": title, "link": link,
                "source_url": source_url, "source_name": source_name,
                "date": date_str,
            })
    except Exception:
        pass
    return results

def ddg_broad_search(keyword, max_results=80):
    if not DDGS:
        return []
    for _ in range(DDG_MAX_RETRIES):
        try:
            with DDGS() as d:
                return list(d.text(keyword, region="tw-tzh", max_results=max_results))
        except Exception as e:
            err = str(e)
            if "429" in err or "ratelimit" in err.lower():
                time.sleep(10)
                continue
            time.sleep(1)
    return []

def ddg_site_search(keyword, domain):
    if not DDGS:
        return []
    q = f"{keyword} site:{domain}"
    for _ in range(DDG_MAX_RETRIES):
        try:
            with DDGS() as d:
                return list(d.text(q, region="tw-tzh", max_results=5))
        except Exception as e:
            err = str(e)
            if "429" in err or "ratelimit" in err.lower():
                time.sleep(10)
                continue
            time.sleep(1)
    return []

# ═══════════════════════════════════════════════════════
# Pre-filter
# ═══════════════════════════════════════════════════════

def is_likely_article(url, title, brand_keywords):
    path = urlparse(url).path.rstrip("/")
    if not path or path == "/":
        return False
    segments = [s for s in path.split("/") if s]
    if len(segments) < 1:
        return False
    title_lower = title.lower()
    skip_url = ['search?', '/tag/', '/category/', '/recruit', '/job/',
                '/stock/', '/finance/quote', '/redirect_file']
    if any(p in url.lower() for p in skip_url):
        return False
    if re.match(r'^[a-z0-9.-]+\.(com|tw|net|org)', title_lower.strip()):
        return False
    if len(title.strip()) < 8:
        return False
    has_keyword = any(kw.lower() in title_lower for kw in brand_keywords if len(kw) > 1)
    if not has_keyword:
        return False
    return True

# ═══════════════════════════════════════════════════════
# LLM verification — 加入日期意識
# ═══════════════════════════════════════════════════════

def call_gemini(prompt):
    if not GEMINI_API_KEY:
        return None
    try:
        resp = httpx.post(
            GEMINI_URL, params={"key": GEMINI_API_KEY},
            json={"contents": [{"parts": [{"text": prompt}]}],
                  "generationConfig": {"temperature": 0, "maxOutputTokens": 8192}},
            timeout=60)
        resp.raise_for_status()
        text = (resp.json().get("candidates", [{}])[0]
                .get("content", {}).get("parts", [{}])[0].get("text", ""))
        text = text.strip()
        if text.startswith("```"):
            text = text.split("\n", 1)[-1].rsplit("```", 1)[0]
        return json.loads(text)
    except Exception as e:
        return None

BRAND_PROMPT_MAP = {
    "全家": {
        "full_name": "全家便利商店（FamilyMart）",
        "short_name": "全家",
        "competitor_note": "其他品牌（7-11、統一）的活動，沒提到全家",
    },
    "7-11": {
        "full_name": "7-ELEVEN（統一超商）",
        "short_name": "7-ELEVEN",
        "competitor_note": "其他品牌（全家、FamilyMart）的活動，沒提到7-ELEVEN或統一超商",
    },
}

def verify_batch(task_name, press_date, candidates, brand="全家"):
    BATCH = 20
    verified = []
    bp = BRAND_PROMPT_MAP.get(brand, BRAND_PROMPT_MAP["全家"])
    for batch_start in range(0, len(candidates), BATCH):
        batch = candidates[batch_start:batch_start + BATCH]
        items_text = ""
        for idx, c in enumerate(batch):
            items_text += (
                f"\n---\n#{idx+1}\n"
                f"標題: {c.get('標題', '')}\n"
                f"媒體: {c.get('媒體', '')}\n"
                f"連結: {c.get('連結', '')}\n"
                f"日期: {c.get('日期', '未知')}\n"
            )
        prompt = (
            f"你是嚴格的新聞露出檢核員。以下是搜尋「{task_name}」（{bp['full_name']}的活動/產品）的候選結果。\n"
            f"本次新聞稿發稿日為：{press_date}\n\n"
            f"請嚴格判斷每則結果是否是一篇**真正報導**{bp['full_name']}「{task_name}」的**新聞文章**。\n\n"
            f"【判斷標準 — 嚴格執行】\n"
            f"✅ 相關：文章主題是在報導{bp['full_name']}的「{task_name}」相關活動或商品\n"
            f"✅ 相關：綜合報導中有明確提及{bp['short_name']}的「{task_name}」（如「超商活動比較」含{bp['short_name']}段落）\n"
            f"❌ 不相關：首頁、分類頁、關鍵字頁、搜尋結果頁（URL通常很短或含 /tag/ /category/）\n"
            f"❌ 不相關：股票/財報/人力資源頁面只是因為含有公司名\n"
            f"❌ 不相關：只提到部分字眼但文章主題完全不同\n"
            f"❌ 不相關：{bp['competitor_note']}\n"
            f"❌ 不相關：文章日期在 {press_date} 之前（即往年同名活動的報導，不是本次新聞稿的露出）\n"
            f"❌ 不相關：日期標示「未知」但標題明顯是往年的舊文（如提及2024、2025年份詞彙、去年）\n"
            f"⚠️ 注意：日期「未知」但內容看起來是 2026 年新聞稿的露出 → 判 true（給予合理懷疑空間）\n\n"
            f"候選列表：{items_text}\n\n"
            f"請回答 JSON array，每個元素格式：\n"
            f'{{"id": 1, "relevant": true/false, "type": "原生"/"轉載", "reason": "簡短理由"}}\n'
            f"預設為 false（不確定就判 false）。只回 JSON array。"
        )
        verdicts = call_gemini(prompt)
        time.sleep(0.3)
        if verdicts is None:
            continue
        verdict_map = {v.get("id"): v for v in verdicts if v.get("id") is not None}
        batch_pass = 0
        for idx, c in enumerate(batch):
            v = verdict_map.get(idx + 1, {})
            if v.get("relevant", False):
                rtype = str(v.get("type", "原生"))
                c["原生/轉載"] = "轉載" if "轉載" in rtype else "原生"
                c["llm_reason"] = v.get("reason", "")
                verified.append(c)
                batch_pass += 1
        console.print(f"    [dim]batch {batch_start+1}-{batch_start+len(batch)}: "
                      f"[green]{batch_pass} pass[/] / [red]{len(batch)-batch_pass} reject[/][/dim]")
    return verified

# ═══════════════════════════════════════════════════════
# HIGH-VALUE media for Phase 1c
# ═══════════════════════════════════════════════════════

TOP_MEDIA = {
    "三立新聞", "自由時報", "民視新聞", "東森新聞", "華視新聞", "台視新聞",
    "TVBS新聞", "中天新聞網", "中央社", "鏡週刊", "鏡新聞",
    "風傳媒", "上報 Up Media", "NewTalk新頭殼", "CNEWS匯流新聞網",
    "今周刊", "遠見", "天下雜誌", "商業週刊_新聞", "數位時代",
    "科技新報", "鉅亨網", "工商時報", "經濟日報",
    "ETtoday旅遊雲", "ETtoday財經雲",
    "中華新聞雲/中華日報", "壹蘋新聞網", "聯合報",
    "公視新聞", "年代新聞", "非凡新聞",
    "卡優新聞網", "聚財網-財經新聞", "豐雲學堂",
    "奧丁丁新聞OwlNews", "PopDaily波波黛莉", "好想出去玩",
    "女人我最大", "哈潑時尚BAZAAR", "ELLE", "Marie claire 美麗佳人",
    "食力foodNEXT", "親子天下", "關鍵評論網",
    "知新聞", "PChome股市", "uniopen新聞", "CMoney 追訊",
    "好新聞", "威傳媒", "桃園電子報", "台灣好新聞報",
    "Styletc 樂時尚", "GIRLSTALK", "好房News",
    "巴哈姆特新聞網", "4gamers", "妞新聞",
    "LIFE生活網", "Taiwan News Agency 台灣新聞通訊社",
    "臺灣時報Taiwan Times", "紅新聞", "台灣新生報",
    "ETtoday新聞雲", "旺得富理財網", "觸Mii", "玩咖Playing",
    "鏡週刊", "NOWnews今日新聞", "KiraKacha去啦",
    "祝你健康", "壹電視新聞台", "理財周刊", "火報",
    "東台灣新聞網", "Money101", "花花日報",
    "COSMOPOLITAN Taiwan", "美力圈", "udn STYLE",
    "BEAUTY美人圈", "WalkerLand 窩客島", "食尚玩家",
    "鏡報新聞網", "太報", "MSN新聞", "民眾網",
}

# ═══════════════════════════════════════════════════════
# Main scan pipeline
# ═══════════════════════════════════════════════════════

def scan_single_task(task_name, config, media_list, domain_map, progress=None, task_id=None):
    keywords = config["keywords"]
    fragments = config.get("press_title_fragments", [])
    brand_kws = config.get("brand_keywords", [])
    press_date = config.get("press_date", "2026-01-01")

    seen_keys = set()
    candidates = []
    date_rejected = 0
    phase_stats = {"GNews": 0, "DDG-broad": 0, "DDG-site": 0, "DDG-title": 0}
    covered_media = set()
    media_name_set = {m["name"] for m in media_list}
    media_name_norm = {m["name"].replace(" ", "").lower(): m["name"] for m in media_list}

    def update_progress(desc):
        if progress and task_id is not None:
            progress.update(task_id, description=f"  [cyan]{task_name}[/] {desc}")

    def add_candidate(media_name, title, link, date_str, source, keyword):
        nonlocal date_rejected
        key = f"{media_name}|{link}"
        if key in seen_keys:
            return False
        seen_keys.add(key)
        after = is_after_press_date(date_str, press_date)
        if after is False:
            date_rejected += 1
            return False
        candidates.append({
            "新聞": task_name, "日期": date_str, "媒體": media_name,
            "標題": title, "連結": link, "關鍵字": keyword, "來源": source,
        })
        covered_media.add(media_name)
        return True

    # ── Phase 1a: Google News RSS (with date filter) ──
    update_progress("Phase 1a: Google News RSS")
    all_queries = list(keywords) + list(fragments)
    for kw in all_queries:
        articles = google_news_rss(kw, after_date=press_date)
        for a in articles:
            media_name = None
            real_url = a.get("source_url", "") or a["link"]
            media_name, _ = match_domain(real_url, domain_map)
            if not media_name:
                media_name, _ = match_domain(a["link"], domain_map)
            if not media_name:
                sn = a.get("source_name", "").strip()
                if sn in media_name_set:
                    media_name = sn
                else:
                    sn_norm = sn.replace(" ", "").lower()
                    media_name = media_name_norm.get(sn_norm)
            if not media_name:
                continue
            title = a["title"].rsplit(" - ", 1)[0].strip()
            if not is_likely_article(a["link"], title, brand_kws):
                continue
            if add_candidate(media_name, title, a["link"], a["date"], "GNews", kw):
                phase_stats["GNews"] += 1
        time.sleep(0.3)
    console.print(f"    Phase 1a: [green]{phase_stats['GNews']}[/] 筆, {len(covered_media)} 媒體 "
                  f"([red]{date_rejected} date-rejected[/])")

    # ── Phase 1b: DDG broad search ──
    update_progress("Phase 1b: DDG 廣域搜尋")
    ddg_kw_list = keywords[:6]
    for kw in ddg_kw_list:
        results = ddg_broad_search(kw, max_results=80)
        for r in results:
            url = r.get("href", "")
            title = r.get("title", "")
            body = r.get("body", "")
            if not url or not title:
                continue
            media_name, _ = match_domain(url, domain_map)
            if not media_name:
                continue
            if not is_likely_article(url, title, brand_kws):
                continue
            date_str = extract_date_from_snippet(body) or extract_date_from_snippet(title)
            if add_candidate(media_name, title, url, date_str, "DDG-broad", kw):
                phase_stats["DDG-broad"] += 1
        time.sleep(SEARCH_DELAY)

    for frag in fragments:
        results = ddg_broad_search(f'"{frag}"', max_results=30)
        for r in results:
            url = r.get("href", "")
            title = r.get("title", "")
            body = r.get("body", "")
            if not url or not title:
                continue
            media_name, _ = match_domain(url, domain_map)
            if not media_name:
                continue
            if not is_likely_article(url, title, brand_kws):
                continue
            date_str = extract_date_from_snippet(body) or extract_date_from_snippet(title)
            if add_candidate(media_name, title, url, date_str, "DDG-title", frag):
                phase_stats["DDG-title"] += 1
        time.sleep(SEARCH_DELAY)
    console.print(f"    Phase 1b: [green]+{phase_stats['DDG-broad']+phase_stats['DDG-title']}[/] 筆, "
                  f"{len(covered_media)} 媒體 ([red]{date_rejected} date-rejected[/])")

    # ── Phase 1c: DDG site search for uncovered high-value media ──
    # Use batched approach: only 1 keyword per media, with global timeout
    update_progress("Phase 1c: DDG site 補漏")
    uncovered_important = [m for m in media_list
                           if m["name"] not in covered_media and m["name"] in TOP_MEDIA]
    phase1c_start = time.time()
    PHASE1C_TIMEOUT = 300
    ddg_consecutive_fails = 0
    for m in uncovered_important:
        if time.time() - phase1c_start > PHASE1C_TIMEOUT:
            console.print(f"    [yellow]Phase 1c timeout ({PHASE1C_TIMEOUT}s), 跳過剩餘[/]")
            break
        if ddg_consecutive_fails >= 5:
            console.print(f"    [yellow]DDG 連續失敗 {ddg_consecutive_fails} 次, 跳過剩餘[/]")
            break
        domain, name = m["domain"], m["name"]
        kw = keywords[0]
        if domain in AGGREGATOR_DOMAINS:
            q_results = ddg_broad_search(f"{kw} {name}", max_results=5)
        else:
            q_results = ddg_site_search(kw, domain)
        if not q_results:
            ddg_consecutive_fails += 1
            time.sleep(0.5)
            continue
        ddg_consecutive_fails = 0
        for r in q_results:
            url = r.get("href", "")
            title = r.get("title", "")
            body = r.get("body", "")
            if not url or not title:
                continue
            rd = urlparse(url).netloc.replace("www.", "")
            if domain not in rd and rd not in domain:
                continue
            if not is_likely_article(url, title, brand_kws):
                continue
            date_str = extract_date_from_snippet(body) or extract_date_from_snippet(title)
            if add_candidate(name, title, url, date_str, "DDG-site", kw):
                phase_stats["DDG-site"] += 1
                break
        time.sleep(0.8)
    console.print(f"    Phase 1c: [green]+{phase_stats['DDG-site']}[/] 筆, "
                  f"{len(covered_media)} 媒體")

    console.print(f"    候選合計: [bold]{len(candidates)}[/] 筆 (date-rejected: [red]{date_rejected}[/])")

    # ── Phase 2: LLM verification ──
    update_progress("Phase 2: LLM 驗證")
    task_brand = config.get("brand", "全家")
    verified = verify_batch(task_name, press_date, candidates, brand=task_brand)
    verified_media = len(set(v["媒體"] for v in verified))
    console.print(f"    LLM 通過: [bold green]{len(verified)}[/] 筆, {verified_media} 媒體")

    return {
        "task": task_name,
        "press_date": press_date,
        "candidates_count": len(candidates),
        "date_rejected": date_rejected,
        "results": verified,
        "stats": phase_stats,
    }


# ═══════════════════════════════════════════════════════
# CLI: scan command
# ═══════════════════════════════════════════════════════

def cmd_scan(args):
    console.print(Panel("[bold]Pandora Scanner v4.1[/] — 掃描模式", style="blue"))

    media_list = load_media_list()
    domain_map = build_domain_map(media_list)
    console.print(f"  媒體清單: [cyan]{len(media_list)}[/] 個, API key: {'[green]✓[/]' if GEMINI_API_KEY else '[red]✗[/]'}")

    brand = getattr(args, 'brand', None) or "全家"
    tasks = TASKS_CONFIG
    if args.task:
        if args.task not in tasks:
            console.print(f"[red]找不到 '{args.task}'[/], 可用: {', '.join(tasks)}")
            return
        tasks = {args.task: tasks[args.task]}
    tasks = {k: v for k, v in tasks.items()
             if v["month"] == args.month and v.get("brand", "全家") == brand}
    console.print(f"  品牌: [bold cyan]{brand}[/], 準備掃描: [cyan]{len(tasks)}[/] 個任務 ({args.month})\n")

    OUTPUT_DIR.mkdir(exist_ok=True)
    all_results = []
    task_summaries = []

    for task_name, cfg in tasks.items():
        console.rule(f"[bold]{task_name}[/] (發稿: {cfg['press_date']})")
        result = scan_single_task(task_name, cfg, media_list, domain_map)
        out_path = OUTPUT_DIR / f"{task_name}_v4.json"
        out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        all_results.extend(result["results"])
        task_summaries.append(result)
        console.print(f"    已存: [dim]{out_path.name}[/dim]\n")

    brand_suffix = f"_{brand}" if brand != "全家" else ""
    combined_path = OUTPUT_DIR / f"v4_combined{brand_suffix}.json"
    combined = {
        "version": "v4.1",
        "brand": brand,
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "results": all_results,
        "total": len(all_results),
        "media_count": len(set(r["媒體"] for r in all_results)),
    }
    combined_path.write_text(json.dumps(combined, ensure_ascii=False, indent=2), encoding="utf-8")

    _print_scan_summary(task_summaries, all_results)

def _print_scan_summary(task_summaries, all_results):
    console.print()
    console.rule("[bold]掃描結果總覽[/]")

    table = Table(box=box.SIMPLE_HEAVY, show_header=True)
    table.add_column("任務", style="cyan", width=16)
    table.add_column("發稿日", style="dim", width=12)
    table.add_column("候選", justify="right", width=6)
    table.add_column("日期淘汰", justify="right", style="red", width=8)
    table.add_column("LLM通過", justify="right", style="green", width=8)
    table.add_column("媒體數", justify="right", width=6)
    table.add_column("GNews", justify="right", width=6)
    table.add_column("DDG", justify="right", width=6)

    total_cand, total_rej, total_pass = 0, 0, 0
    for s in task_summaries:
        n_pass = len(s["results"])
        n_media = len(set(r["媒體"] for r in s["results"]))
        ddg = s["stats"]["DDG-broad"] + s["stats"]["DDG-site"] + s["stats"]["DDG-title"]
        table.add_row(
            s["task"], s["press_date"],
            str(s["candidates_count"]), str(s["date_rejected"]),
            str(n_pass), str(n_media),
            str(s["stats"]["GNews"]), str(ddg))
        total_cand += s["candidates_count"]
        total_rej += s["date_rejected"]
        total_pass += n_pass
    table.add_section()
    table.add_row(
        "[bold]合計[/]", "",
        f"[bold]{total_cand}[/]", f"[bold]{total_rej}[/]",
        f"[bold]{total_pass}[/]",
        f"[bold]{len(set(r['媒體'] for r in all_results))}[/]", "", "")
    console.print(table)

    src = Counter(r.get("來源", "?") for r in all_results)
    console.print("\n  [bold]來源分佈:[/]")
    for s, c in src.most_common():
        console.print(f"    {s}: [cyan]{c}[/]")

    date_ok = sum(1 for r in all_results if r.get("日期"))
    date_empty = sum(1 for r in all_results if not r.get("日期"))
    console.print(f"\n  [bold]日期:[/] 有日期 [green]{date_ok}[/], 無日期 [yellow]{date_empty}[/]")

    if all_results:
        dates = sorted([r["日期"] for r in all_results if r.get("日期")])
        if dates:
            console.print(f"    日期範圍: {dates[0]} ~ {dates[-1]}")


# ═══════════════════════════════════════════════════════
# CLI: report command — 比較 v4 vs 舊版 vs 意藍
# ═══════════════════════════════════════════════════════

def cmd_report(args):
    console.print(Panel("[bold]Pandora Scanner v4.1[/] — 比較報告", style="green"))

    combined_path = OUTPUT_DIR / "v4_combined.json"
    if not combined_path.exists():
        console.print("[red]尚無 v4 結果，請先執行 scan[/]")
        return

    with open(combined_path) as f:
        v4 = json.load(f)
    v4_results = v4["results"]
    console.print(f"  v4 結果: [cyan]{len(v4_results)}[/] 筆, {v4.get('media_count', '?')} 媒體")

    # Load old results
    old_results = _load_old_results()
    console.print(f"  舊版結果: [cyan]{len(old_results)}[/] 筆")

    # Load 意藍
    yilan_media, yilan_total = _load_yilan_summary()
    console.print(f"  意藍報告: [cyan]{yilan_total}[/] 篇, {len(yilan_media)} 媒體\n")

    # ── Main comparison table ──
    table = Table(title="v4 vs 舊版 vs 意藍", box=box.DOUBLE_EDGE, show_header=True)
    table.add_column("指標", style="bold", width=28)
    table.add_column("v4.1 新版", justify="right", style="green", width=12)
    table.add_column("舊版(合併)", justify="right", style="yellow", width=12)
    table.add_column("意藍", justify="right", style="cyan", width=12)

    v4_media = set(r["媒體"] for r in v4_results)
    old_media = set(_get(r, "媒體") for r in old_results if _get(r, "媒體"))
    yilan_names = set(yilan_media.keys())

    def _norm(s):
        return s.replace(" ", "").replace("　", "").lower()

    def _fuzzy_cover(our_media, yilan_set):
        covered = set()
        our_norm = {_norm(m): m for m in our_media}
        for ym in yilan_set:
            if ym in our_media or _norm(ym) in our_norm:
                covered.add(ym)
        return covered

    v4_yilan_cover = _fuzzy_cover(v4_media, yilan_names)
    old_yilan_cover = _fuzzy_cover(old_media, yilan_names)

    table.add_row("總結果數", str(len(v4_results)), str(len(old_results)), str(yilan_total))
    table.add_row("命中媒體數", str(len(v4_media)), str(len(old_media)), str(len(yilan_names)))
    table.add_row("覆蓋意藍媒體數", str(len(v4_yilan_cover)), str(len(old_yilan_cover)), "—")
    table.add_row("覆蓋意藍媒體%",
                  f"{len(v4_yilan_cover)*100//max(1,len(yilan_names))}%",
                  f"{len(old_yilan_cover)*100//max(1,len(yilan_names))}%", "—")

    # Date quality
    v4_has_date = sum(1 for r in v4_results if r.get("日期"))
    v4_no_date = len(v4_results) - v4_has_date
    if v4_results:
        old_dates = [r["日期"] for r in v4_results if r.get("日期") and r["日期"] < "2026-01-01"]
        table.add_row("有日期比例", f"{v4_has_date}/{len(v4_results)}", "—", "—")
        table.add_row("2026前舊文", f"[{'red' if old_dates else 'green'}]{len(old_dates)}[/]", "—", "—")
    console.print(table)

    # ── Per-task comparison ──
    console.print()
    task_table = Table(title="各任務結果數", box=box.SIMPLE_HEAVY)
    task_table.add_column("任務", style="cyan", width=16)
    task_table.add_column("v4.1", justify="right", style="green", width=8)
    task_table.add_column("舊版", justify="right", style="yellow", width=8)

    v4_by_task = Counter(r["新聞"] for r in v4_results)
    old_by_task = Counter(_get(r, "新聞") or "?" for r in old_results)
    for task in sorted(set(list(v4_by_task.keys()) + list(old_by_task.keys()))):
        task_table.add_row(task, str(v4_by_task.get(task, 0)), str(old_by_task.get(task, 0)))
    console.print(task_table)

    # ── Media coverage diff ──
    new_coverage = v4_yilan_cover - old_yilan_cover
    if new_coverage:
        console.print(f"\n  [bold green]v4 新增覆蓋的意藍媒體（舊版沒有）: {len(new_coverage)} 家[/]")
        for m in sorted(new_coverage, key=lambda x: yilan_media.get(x, 0), reverse=True):
            console.print(f"    ✅ {m}: 意藍 {yilan_media.get(m, 0)} 則")

    still_missing = yilan_names - v4_yilan_cover
    if still_missing:
        console.print(f"\n  [bold red]v4 仍漏掉的意藍媒體 Top 10:[/]")
        sorted_missing = sorted(still_missing, key=lambda x: yilan_media.get(x, 0), reverse=True)
        for m in sorted_missing[:10]:
            console.print(f"    ❌ {m}: 意藍 {yilan_media.get(m, 0)} 則")

    # Source distribution
    src = Counter(r.get("來源", "?") for r in v4_results)
    console.print(f"\n  [bold]v4 來源分佈:[/]")
    for s, c in src.most_common():
        console.print(f"    {s}: [cyan]{c}[/]")


def _get(row, zh_key):
    """Get a value by Chinese key, falling back to English equivalents."""
    FIELD_MAP = {
        "媒體": ["媒體", "media"],
        "新聞": ["新聞", "task"],
        "標題": ["標題", "title"],
        "日期": ["日期", "date"],
        "連結": ["連結", "link"],
        "來源": ["來源", "source"],
    }
    for k in FIELD_MAP.get(zh_key, [zh_key]):
        if k in row:
            return row[k]
    return None

def _load_old_results():
    merged = OLD_RESULTS_DIR / "一月完整合併.json"
    if merged.exists():
        try:
            data = json.loads(merged.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return data
            return data.get("results", data.get("data", []))
        except Exception:
            pass
    results = []
    for f in OLD_RESULTS_DIR.glob("*.json"):
        if "完整合併" in f.name:
            continue
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                items = data.get("results", data.get("data", []))
            elif isinstance(data, list):
                items = data
            else:
                continue
            results.extend(items)
        except Exception:
            continue
    return results

def _load_yilan_summary():
    media_counts = Counter()
    total = 0

    # Try loading from cached JSON first (created by previous analysis)
    cache_path = OUTPUT_DIR / "yilan_cache.json"
    if cache_path.exists():
        try:
            data = json.loads(cache_path.read_text(encoding="utf-8"))
            return data.get("media_counts", {}), data.get("total", 0)
        except Exception:
            pass

    try:
        sys.path.insert(0, os.path.expanduser(
            "~/openclaw/skills/pandora-news/venv/lib/python3.9/site-packages"))
        from openpyxl import load_workbook
        xlsx_files = list(YILAN_DIR.glob("*.xlsx"))
        if not xlsx_files:
            import subprocess
            result = subprocess.run(["ls", str(YILAN_DIR)], capture_output=True, text=True)
            xlsx_files = [YILAN_DIR / f for f in result.stdout.strip().split("\n")
                         if f.endswith(".xlsx") and not f.startswith("~")]

        for xlsx in xlsx_files:
            if not xlsx.exists() or xlsx.name.startswith("~"):
                continue
            try:
                wb = load_workbook(str(xlsx), read_only=True, data_only=True)
            except Exception:
                continue
            for ws in wb.worksheets:
                if "新聞列表" not in ws.title and "list" not in ws.title.lower():
                    continue
                header = None
                for row in ws.iter_rows(values_only=True):
                    if header is None:
                        cells = [str(c or "").strip() for c in row]
                        if any("媒體" in c for c in cells):
                            header = cells
                            media_col = next(i for i, c in enumerate(cells) if "媒體" in c)
                            continue
                        continue
                    vals = list(row)
                    if len(vals) > media_col and vals[media_col]:
                        mname = str(vals[media_col]).strip()
                        if mname:
                            media_counts[mname] += 1
                            total += 1
            wb.close()
    except Exception as e:
        console.print(f"  [yellow]意藍讀取警告: {e}[/]")

    if media_counts:
        cache_path.parent.mkdir(exist_ok=True)
        cache_path.write_text(json.dumps(
            {"media_counts": dict(media_counts), "total": total},
            ensure_ascii=False, indent=2), encoding="utf-8")

    return dict(media_counts), total


# ═══════════════════════════════════════════════════════
# CLI: results command — 查看/抽查結果
# ═══════════════════════════════════════════════════════

def cmd_results(args):
    console.print(Panel("[bold]Pandora Scanner v4.1[/] — 結果檢視", style="magenta"))

    combined_path = OUTPUT_DIR / "v4_combined.json"
    if not combined_path.exists():
        console.print("[red]尚無結果[/]")
        return

    with open(combined_path) as f:
        v4 = json.load(f)
    results = v4["results"]

    if args.task:
        results = [r for r in results if r["新聞"] == args.task]
        console.print(f"  篩選任務: [cyan]{args.task}[/] → {len(results)} 筆")

    if args.check_quality:
        _quality_check(results)
        return

    table = Table(box=box.ROUNDED, show_lines=True, width=120)
    table.add_column("#", style="dim", width=4)
    table.add_column("日期", width=10)
    table.add_column("媒體", style="cyan", width=18)
    table.add_column("標題", width=55)
    table.add_column("來源", style="dim", width=9)
    table.add_column("類型", width=5)

    limit = args.limit or 30
    for i, r in enumerate(results[:limit]):
        date_style = "green" if r.get("日期", "").startswith("2026") else "yellow" if r.get("日期") else "red"
        table.add_row(
            str(i+1),
            f"[{date_style}]{r.get('日期', '??')}[/]",
            r.get("媒體", ""),
            r.get("標題", "")[:55],
            r.get("來源", ""),
            r.get("原生/轉載", "?")
        )
    console.print(table)
    if len(results) > limit:
        console.print(f"  [dim]顯示前 {limit} 筆，共 {len(results)} 筆 (用 --limit N 看更多)[/dim]")

def _quality_check(results):
    console.print("\n[bold]品質檢核[/]")
    issues = {"no_date": [], "old_date": [], "suspicious_title": []}

    for r in results:
        d = r.get("日期", "")
        if not d:
            issues["no_date"].append(r)
        elif d < "2026-01-01":
            issues["old_date"].append(r)

        title = r.get("標題", "")
        if re.search(r'2024|2025年|去年', title):
            issues["suspicious_title"].append(r)

    console.print(f"  無日期: [yellow]{len(issues['no_date'])}[/]")
    console.print(f"  2026前舊文: [red]{len(issues['old_date'])}[/]")
    console.print(f"  標題含舊年份: [red]{len(issues['suspicious_title'])}[/]")

    if issues["old_date"]:
        console.print("\n  [red]舊文章 (前10):[/]")
        for r in issues["old_date"][:10]:
            console.print(f"    [{r['日期']}] {r['媒體']}: {r['標題'][:60]}")

    if issues["suspicious_title"]:
        console.print("\n  [red]標題含舊年份:[/]")
        for r in issues["suspicious_title"][:10]:
            console.print(f"    [{r.get('日期', '??')}] {r['媒體']}: {r['標題'][:60]}")


# ═══════════════════════════════════════════════════════
# CLI: press command — 查看/管理新聞稿
# ═══════════════════════════════════════════════════════

def cmd_press(args):
    console.print(Panel("[bold]Pandora Scanner v4.1[/] — 新聞稿管理", style="yellow"))

    table = Table(box=box.SIMPLE_HEAVY, show_header=True)
    table.add_column("任務", style="cyan", width=16)
    table.add_column("發稿日", width=12)
    table.add_column("關鍵字數", justify="right", width=8)
    table.add_column("品牌詞", width=30)
    table.add_column("標題片段", width=40)

    for name, cfg in sorted(TASKS_CONFIG.items(), key=lambda x: x[1].get("press_date", "")):
        table.add_row(
            name, cfg.get("press_date", "?"),
            str(len(cfg.get("keywords", []))),
            ", ".join(cfg.get("brand_keywords", [])[:4]),
            cfg.get("press_title_fragments", [""])[0][:40] if cfg.get("press_title_fragments") else "—"
        )
    console.print(table)

    if args.task:
        cfg = TASKS_CONFIG.get(args.task)
        if cfg:
            console.print(f"\n[bold]{args.task}[/] 詳細設定:")
            console.print(f"  發稿日: {cfg['press_date']}")
            console.print(f"  關鍵字:")
            for kw in cfg["keywords"]:
                console.print(f"    • {kw}")
            console.print(f"  品牌詞: {', '.join(cfg['brand_keywords'])}")
            console.print(f"  新聞稿標題:")
            for f in cfg.get("press_title_fragments", []):
                console.print(f"    • {f}")


# ═══════════════════════════════════════════════════════
# CLI: monthly command — 一口氣產出完整月報 → Google Sheet
# ═══════════════════════════════════════════════════════

SOCIAL_KEYWORDS = [
    "facebook", "粉絲團", "instagram", "threads", "tiktok",
    "youtube", "ptt", "dcard", "mobile01", "plurk", "blogger",
    "痞客邦", "medium.com", "popdaily", "pixnet",
]

AGGREGATOR_MEDIA = [
    "yahoo新聞", "yahoo股市", "line today", "msn", "msn新聞",
    "pchome online 新聞", "pchome股市", "蕃新聞", "match生活網",
    "翻爆",
]

PRESS_RELEASE_TITLES = {
    "全家草莓季": "「全家」草莓季17粉嫩登場！攜手日本高人氣IP「ASAMIMICHAN」萌翻全台",
    "全家UCC咖啡": "「全家」再攜UCC推雙冠軍監製新品 Let's Café 阿里山極選綜合咖啡65元開喝",
    "全家開運鮮食": "「全家」預熱年味！紅運烏魚子紹興飯糰等6款開運鮮食登場",
    "全家年菜預購": "「全家」搶攻圍爐商機！2026金馬年菜預購推百款星級年菜FamiPort一站購足",
    "全家蜷川實花": "粉絲必收！「全家」攜手蜷川實花展推獨家限定杯身與杯套再享展覽優惠",
    "全家特力屋": "瞄準除舊佈新商機！「全家」化身神隊友 攜特力屋齊推「居家微整型」煥新提案",
    "全家溏心蛋": "「全家」熟食區新蛋報到！首推「用撈的」日式溏心蛋 25元爆汁開吃",
    "全家高山茶": "寒流飄茶香！高山茶進駐「全家」Let's Tea現煮精品「蘭韻梨山烏龍」49元開喝",
    "全家超人力霸王": "光之巨人降臨！「全家」×2026高雄冬日遊樂園 超人力霸王聯名杯身登場",
    "全家寒流抗寒": "強烈冷氣團明來襲 「全家」熱飲、熱食、暖心織品 抗寒限定好康報到！",
    "全家伴手禮": "「全家」搶返鄉商機！實體店開賣最強伴手禮「陳耀訓．紅土蛋黃酥禮盒」",
    "711清水服務區週年慶": "統一超商經營清水服務區1週年慶！首公開懂買、懂吃、懂拍旅遊攻略",
    "711米其林法餐": "免訂位享星級儀式感！7-ELEVEN米其林30顆星主廚監製法餐上桌",
    "711年節社交經濟": "7-ELEVEN全面搶攻年節「社交經濟」！開運年菜、尾牙家電馬上登場",
    "711把愛找回來公益": "雙向奔赴的暖心公益！2026年7-ELEVEN把愛找回來公益募款平台擴大合作130個公益團體",
    "711智能果昔機": "7-ELEVEN首度導入「智能果昔機」，獨家「鮮玉米濃湯」現打熱熱喝",
    "711阜杭豆漿飯糰": "7-ELEVEN「阜杭豆漿飯糰」締造億元台式小吃傳奇！尾牙開運、學測考生必吃",
    "711小熊維尼集點": "小熊維尼滿100周年了！7-ELEVEN全店集點推出80款獨家療癒新品",
    "711馬年前哨戰優惠": "0115優惠資訊（馬年前哨戰）",
    "711Fresh橋港門市": "卡位八里淡水生活圈，7-ELEVEN雙北首間「Fresh橋港門市」來啦！",
    "711金馬開運活動": "1/21~1/25 7-ELEVEN金馬開運好運來活動",
    "711藝伎咖啡": "「阿里山豆御香藝伎咖啡」每杯2,000元限定門市限量開賣",
    "711抗寒保暖": "0120優惠資訊（抗寒保暖對策）",
    "711金馬年開運": "新年到「馬上開心」7-ELEVEN多重回饋迎金馬年 超過18項開運美食好運一起發！",
    "711東南亞美食": "吃「東南亞美食」來7-ELEVEN，正餐、零食到水果帶來道地家鄉味",
    "711桂氣茶飲優惠": "0126優惠資訊（開運「桂」氣茶飲手搖控開喝）",
    "711西村優志集點": "7-ELEVEN全店精品集點首度攜手日本人氣創作者「西村優志」",
    "711開運福袋": "7-ELEVEN第二波14款開運福袋攜手7大超萌IP肖像 2月11日起限量開賣",
    "711香菜美食": "就愛這一味！7-ELEVEN逾20款香菜美食等你來「開香」",
    "711OPEN家族貼圖": "社群最新拜年神器 OPEN!家族2026年16款全新免費貼圖2月3日開放下載",
    "711貓福珊迪": "喵～貓之日 7-ELEVEN聯名「貓福珊迪」甜蜜賣萌！",
    "711開工狂撒幣A": "7-ELEVEN迎新年狂撒幣快來門市淘金，多重優惠回饋總價值達5千萬，開工開學推CITY系列優惠最殺5折起，年後飲控逾百款超值早餐、鮮食組合",
    "711開工狂撒幣B": "7-ELEVEN迎新年狂撒幣快來門市淘金，多重優惠回饋總價值達5千萬，開工開學推CITY系列優惠最殺5折起，年後飲控逾百款超值早餐、鮮食組合",
    "全家減糖無糖飲品": "「全家」響應國健署減糖指標 攜手營養師公會全國聯合會推廣無糖飲品",
}

MONTHLY_ORDER = {
    "7-11": {
        "一月": [
            "711清水服務區週年慶",
            "711米其林法餐",
            "711年節社交經濟",
            "711把愛找回來公益",
            "711智能果昔機",
            "711阜杭豆漿飯糰",
            "711小熊維尼集點",
            "711馬年前哨戰優惠",
            "711Fresh橋港門市",
            "711藝伎咖啡",
            "711抗寒保暖",
            "711金馬年開運",
            "711東南亞美食",
            "711桂氣茶飲優惠",
            "711西村優志集點",
            "711開運福袋",
            "711OPEN家族貼圖",
            "711香菜美食",
            "711貓福珊迪",
        ],
        "二月": [
            "711開工狂撒幣A",
            "711開工狂撒幣B",
        ],
    },
    "全家": {
        "二月": [
            "全家減糖無糖飲品",
        ],
    },
}

def _is_social(article):
    media = (article.get("媒體", "") or "").lower()
    link = (article.get("連結", "") or "").lower()
    combined = media + " " + link
    return any(kw in combined for kw in SOCIAL_KEYWORDS)

def _is_repost(article):
    media = (article.get("媒體", "") or "").lower()
    return any(agg in media for agg in AGGREGATOR_MEDIA)

def _get_sheets_service():
    gsheets_dir = os.path.expanduser("~/openclaw/skills/google-sheets")
    sys.path.insert(0, os.path.join(gsheets_dir, "scripts"))
    token_path = os.path.join(gsheets_dir, "token.json")
    scopes = ['https://www.googleapis.com/auth/spreadsheets']

    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build as api_build

    creds = None
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, scopes)
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
        with open(token_path, 'w') as f:
            f.write(creds.to_json())
    if not creds or not creds.valid:
        raise RuntimeError(f"Google OAuth token 無效或不存在: {token_path}")
    return api_build('sheets', 'v4', credentials=creds)


def cmd_monthly(args):
    console.print(Panel("[bold]Pandora Scanner v4.1[/] — 月報產出", style="bold magenta"))

    brand = getattr(args, 'brand', None) or "全家"
    brand_suffix = f"_{brand}" if brand != "全家" else ""
    combined_path = OUTPUT_DIR / f"v4_combined{brand_suffix}.json"
    if not combined_path.exists():
        console.print(f"[red]尚無 {brand} 掃描結果。請先執行: scanner_v4_test.py scan --brand {brand}[/]")
        return

    with open(combined_path) as f:
        v4 = json.load(f)
    if "results" in v4:
        all_results = v4["results"]
    else:
        all_results = []
        for task_name_key, task_data in v4.items():
            for r in task_data.get("results", []):
                r.setdefault("新聞", task_name_key)
                all_results.append(r)
    all_media = set(r.get("媒體", "") for r in all_results)
    console.print(f"  品牌: [bold cyan]{brand}[/]")
    console.print(f"  載入掃描結果: [cyan]{len(all_results)}[/] 筆, {len(all_media)} 媒體")

    month_label = args.month or "一月"
    tasks_for_month = {k: v for k, v in TASKS_CONFIG.items()
                       if v["month"] == month_label and v.get("brand", "全家") == brand}

    # ── Classify and build report data ──
    task_stats = []
    detail_rows = []
    idx = 1

    ordered_tasks = MONTHLY_ORDER.get(brand, {}).get(month_label)
    if ordered_tasks:
        task_names = [t for t in ordered_tasks if t in tasks_for_month]
    else:
        task_names = sorted(tasks_for_month.keys(),
                            key=lambda t: tasks_for_month[t].get("press_date", ""))

    for task_name in task_names:
        cfg = tasks_for_month[task_name]
        press_date = cfg["press_date"]
        press_title = PRESS_RELEASE_TITLES.get(task_name, task_name)

        task_articles = [r for r in all_results if r.get("新聞") == task_name]
        all_channels = len(set(a["媒體"] for a in task_articles))

        task_stats.append({
            "title": press_title,
            "date": press_date,
            "task_name": task_name,
            "news_ch": all_channels,
            "news_art": len(task_articles),
            "social_ch": 0,
            "social_posts": 0,
        })

        for a in sorted(task_articles, key=lambda x: x.get("日期", "") or "9999"):
            native_or_repost = a.get("原生/轉載", "")
            if not native_or_repost:
                native_or_repost = "轉載" if _is_repost(a) else "原生"
            detail_rows.append({
                "idx": idx,
                "date": a.get("日期", ""),
                "topic": press_title,
                "title": a.get("標題", ""),
                "media": a.get("媒體", ""),
                "link": a.get("連結", ""),
                "author": a.get("作者", "") or a.get("媒體", ""),
                "native_or_repost": native_or_repost,
            })
            idx += 1

    # ── Display CLI summary ──
    console.rule(f"[bold]{month_label}月報總覽[/]")

    stats_table = Table(title="統計報表", box=box.DOUBLE_EDGE, show_header=True)
    stats_table.add_column("新聞標題", width=50)
    stats_table.add_column("發稿日", width=10)
    stats_table.add_column("新聞頻道數", justify="right", style="cyan", width=8)
    stats_table.add_column("新聞總主文數", justify="right", style="green", width=8)

    total_nch, total_nart = 0, 0
    for s in task_stats:
        stats_table.add_row(
            s["title"][:50], s["date"],
            str(s["news_ch"]), str(s["news_art"]))
        total_nch += s["news_ch"]
        total_nart += s["news_art"]

    stats_table.add_section()
    stats_table.add_row("[bold]合計[/]", "",
                        f"[bold]{total_nch}[/]", f"[bold]{total_nart}[/]")
    console.print(stats_table)
    console.print(f"\n  新聞明細: [cyan]{len(detail_rows)}[/] 筆")

    # ── Export to Google Sheet ──
    if args.no_sheet:
        console.print("\n  [dim]--no-sheet: 跳過 Google Sheet 輸出[/]")
        _save_monthly_json(month_label, task_stats, detail_rows, brand)
        return

    console.print("\n[bold]輸出到 Google Sheet...[/]")
    try:
        service = _get_sheets_service()
    except Exception as e:
        console.print(f"  [red]Google Sheets 連線失敗: {e}[/]")
        console.print("  [yellow]改為存 JSON 檔[/]")
        _save_monthly_json(month_label, task_stats, detail_rows, brand)
        return

    sheet_title = f"{brand} {month_label} 月報 ({datetime.now().strftime('%Y-%m-%d')})"
    try:
        result = service.spreadsheets().create(body={
            'properties': {'title': sheet_title},
            'sheets': [
                {'properties': {'title': '統計報表'}},
                {'properties': {'title': '新聞明細'}},
            ]
        }).execute()
        sheet_id = result['spreadsheetId']
        sheet_url = result['spreadsheetUrl']
        console.print(f"  建立 Sheet: [green]{sheet_title}[/]")
    except Exception as e:
        console.print(f"  [red]建立 Sheet 失敗: {e}[/]")
        _save_monthly_json(month_label, task_stats, detail_rows, brand)
        return

    # Write 統計報表
    stats_header = ['新聞標題', '發稿日期', '發稿編號',
                    '新聞頻道數', '新聞總主文數']
    stats_rows = []
    for i, s in enumerate(task_stats, 1):
        stats_rows.append([
            s["title"], s["date"], str(i),
            s["news_ch"], s["news_art"],
        ])
    stats_rows.append([
        '合計', '', '',
        total_nch, total_nart,
    ])

    service.spreadsheets().values().update(
        spreadsheetId=sheet_id,
        range='統計報表!A1',
        valueInputOption='USER_ENTERED',
        body={'values': [stats_header] + stats_rows}
    ).execute()
    console.print(f"  ✓ 統計報表: {len(task_stats)} 篇新聞稿 + 合計")

    # Write 新聞明細
    detail_header = ['編號', '發布時間', '主題', '新聞標題', '網站', '連結', '作者', '原生/轉載']
    detail_data = []
    for d in detail_rows:
        detail_data.append([
            d["idx"], d["date"], d["topic"],
            d["title"], d["media"], d["link"],
            d["author"], d["native_or_repost"],
        ])

    service.spreadsheets().values().update(
        spreadsheetId=sheet_id,
        range='新聞明細!A1',
        valueInputOption='USER_ENTERED',
        body={'values': [detail_header] + detail_data}
    ).execute()
    console.print(f"  ✓ 新聞明細: {len(detail_rows)} 筆")

    # ── Format the sheet ──
    try:
        _format_sheet(service, sheet_id, result)
    except Exception:
        pass

    console.print(f"\n  [bold green]月報已產出！[/]")
    console.print(f"  [link={sheet_url}]{sheet_url}[/link]")
    _save_monthly_json(month_label, task_stats, detail_rows, brand)


def _format_sheet(service, sheet_id, create_result):
    """Apply basic formatting to the sheet."""
    sheets_meta = create_result.get('sheets', [])
    requests = []
    for sheet in sheets_meta:
        sid = sheet['properties']['sheetId']
        requests.append({
            'repeatCell': {
                'range': {'sheetId': sid, 'startRowIndex': 0, 'endRowIndex': 1},
                'cell': {
                    'userEnteredFormat': {
                        'backgroundColor': {'red': 0.2, 'green': 0.2, 'blue': 0.4},
                        'textFormat': {'bold': True, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}},
                    }
                },
                'fields': 'userEnteredFormat(backgroundColor,textFormat)',
            }
        })
        requests.append({
            'updateSheetProperties': {
                'properties': {'sheetId': sid, 'gridProperties': {'frozenRowCount': 1}},
                'fields': 'gridProperties.frozenRowCount',
            }
        })
    if requests:
        service.spreadsheets().batchUpdate(
            spreadsheetId=sheet_id,
            body={'requests': requests}
        ).execute()


def _save_monthly_json(month_label, task_stats, detail_rows, brand="全家"):
    brand_suffix = f"_{brand}" if brand != "全家" else ""
    out_path = OUTPUT_DIR / f"monthly_{month_label}{brand_suffix}.json"
    data = {
        "month": month_label,
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "stats": task_stats,
        "details": detail_rows,
        "total_news": sum(s["news_art"] for s in task_stats),
        "total_social": sum(s["social_posts"] for s in task_stats),
    }
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    console.print(f"  JSON 備份: [dim]{out_path}[/dim]")


# ═══════════════════════════════════════════════════════
# Main CLI entry
# ═══════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Pandora Scanner v4.1 — CLI 新聞露出掃描器",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
指令:
  scan      執行新聞掃描
  monthly   產出完整月報 → Google Sheet（統計報表 + 新聞明細）
  report    產出比較報告 (v4 vs 舊版 vs 意藍)
  results   查看/檢核掃描結果
  press     查看/管理新聞稿設定
        """)
    sub = parser.add_subparsers(dest="command")

    p_scan = sub.add_parser("scan", help="執行掃描")
    p_scan.add_argument("--task", help="指定單一任務")
    p_scan.add_argument("--month", default="一月", help="月份 (預設: 一月)")
    p_scan.add_argument("--brand", default="全家", help="品牌 (全家 / 7-11)")

    p_monthly = sub.add_parser("monthly", help="產出完整月報 → Google Sheet")
    p_monthly.add_argument("--month", default="一月", help="月份 (預設: 一月)")
    p_monthly.add_argument("--brand", default="全家", help="品牌 (全家 / 7-11)")
    p_monthly.add_argument("--no-sheet", action="store_true", help="不輸出 Google Sheet，只存 JSON")

    p_report = sub.add_parser("report", help="比較報告")

    p_results = sub.add_parser("results", help="查看結果")
    p_results.add_argument("--task", help="篩選任務")
    p_results.add_argument("--check-quality", action="store_true", help="品質檢核")
    p_results.add_argument("--limit", type=int, help="顯示筆數")

    p_press = sub.add_parser("press", help="新聞稿管理")
    p_press.add_argument("--task", help="指定任務查詳情")

    args = parser.parse_args()

    console.print()
    console.print("[bold blue]╔══════════════════════════════════════╗[/]")
    console.print("[bold blue]║   Pandora News Scanner v4.1  CLI    ║[/]")
    console.print("[bold blue]╚══════════════════════════════════════╝[/]")
    console.print()

    if args.command == "scan":
        cmd_scan(args)
    elif args.command == "monthly":
        cmd_monthly(args)
    elif args.command == "report":
        cmd_report(args)
    elif args.command == "results":
        cmd_results(args)
    elif args.command == "press":
        cmd_press(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
